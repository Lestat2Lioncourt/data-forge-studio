"""
Import/Export Mixin - JSON import/export of database connections.
"""

from __future__ import annotations

import logging
from typing import TYPE_CHECKING

from PySide6.QtWidgets import QFileDialog

from ...widgets.dialog_helper import DialogHelper
from ...core.i18n_bridge import tr
from ....utils.workspace_export import (
    export_connections_to_json, save_export_to_file, get_export_summary,
    load_import_from_file, import_connections_from_json
)

if TYPE_CHECKING:
    from ....database.config_db import DatabaseConnection

logger = logging.getLogger(__name__)


class DatabaseImportExportMixin:
    """Mixin providing import/export functionality for database connections."""

    # ------------------------------------------------------------------
    # Global multi-result export (one .xlsx with one sheet per result grid
    # across every open query tab in this database manager).
    # ------------------------------------------------------------------
    def _export_all_results_excel(self, tab_widgets=None, parent=None):
        """Build a single .xlsx workbook containing one sheet per non-empty
        result grid across the supplied query-tab widgets. Each sheet reflects
        the user-visible view (filters + sort applied).

        Args:
            tab_widgets: Optional list of QTabWidget to iterate. If None, falls
                back to `self._get_all_tab_widgets()` (database manager scope).
                The workspace passes its own [self.tab_widget] so the export
                is scoped to that workspace's results.
            parent: Optional QWidget used as parent for dialogs / file picker.

        Default filename: first SavedQuery category found among open tabs,
        falling back to "data_export_<YYYYMMDD>".
        Folder: persisted across sessions in user_preferences.export_last_folder.
        """
        from datetime import datetime
        from pathlib import Path
        import re
        from PySide6.QtWidgets import QApplication
        from PySide6.QtCore import Qt
        from PySide6.QtGui import QCursor

        from ..query_tab import QueryTab

        # 1) Collect (sheet_name, dataframe) for every non-empty result grid
        sheets = []
        used_names: set = set()
        truncated_sheets: list = []

        def _sanitize_sheet_name(name: str) -> str:
            cleaned = re.sub(r"[\[\]\\/?*:]", "_", str(name or "")).strip() or "Sheet"
            cleaned = cleaned[:31]
            base = cleaned
            counter = 2
            while cleaned.lower() in used_names:
                suffix = f"_{counter}"
                cleaned = (base[: 31 - len(suffix)]) + suffix
                counter += 1
            used_names.add(cleaned.lower())
            return cleaned

        first_category = None
        EXCEL_MAX_ROWS = 1_048_575

        # Resolve which tab widgets we iterate. Default = database manager scope.
        if tab_widgets is None:
            tab_widgets = self._get_all_tab_widgets()
        dialog_parent = parent if parent is not None else self

        for tab_widget in tab_widgets:
            for i in range(tab_widget.count()):
                widget = tab_widget.widget(i)
                if not isinstance(widget, QueryTab):
                    continue
                # Pick up the category from the first saved-query-linked tab
                if first_category is None:
                    saved_q = getattr(widget, "_saved_query", None)
                    if saved_q is not None and getattr(saved_q, "category", None):
                        first_category = saved_q.category

                result_tabs = getattr(widget, "_result_tabs", None) or []
                results_widget = getattr(widget, "results_tab_widget", None)
                for state in result_tabs:
                    grid = getattr(state, "grid", None)
                    if grid is None:
                        continue
                    df = grid.get_displayed_dataframe()
                    if df is None or len(df) == 0:
                        continue

                    # Find the result tab's display name in its parent QTabWidget
                    raw_name = "Result"
                    if results_widget is not None:
                        idx = results_widget.indexOf(grid)
                        if idx >= 0:
                            raw_name = results_widget.tabText(idx) or raw_name

                    sheet_name = _sanitize_sheet_name(raw_name)

                    if len(df) > EXCEL_MAX_ROWS:
                        truncated_sheets.append((sheet_name, len(df)))
                        df = df.head(EXCEL_MAX_ROWS)
                    sheets.append((sheet_name, df))

        if not sheets:
            DialogHelper.info(tr("export_all_no_results"), tr("export_all_title"), dialog_parent)
            return

        # 2) Build suggested file path (last folder + first-category name)
        try:
            from ....config.user_preferences import UserPreferences
            last_folder = UserPreferences.instance().get("export_last_folder", "") or ""
        except Exception:
            last_folder = ""

        if first_category and first_category != "No category":
            base_name = re.sub(r'[<>:"/\\|?*\x00-\x1f]', "_", first_category).strip() or "data_export"
        else:
            base_name = f"data_export_{datetime.now().strftime('%Y%m%d')}"
        suggested_name = base_name + ".xlsx"
        if last_folder and Path(last_folder).is_dir():
            suggested = str(Path(last_folder) / suggested_name)
        else:
            suggested = suggested_name

        file_path, _ = QFileDialog.getSaveFileName(
            dialog_parent, tr("export_all_title"), suggested, "Excel Files (*.xlsx);;All Files (*)"
        )
        if not file_path:
            return
        if not file_path.lower().endswith(".xlsx"):
            file_path += ".xlsx"

        # Persist the chosen folder for next time (same key as single-grid export)
        try:
            from ....config.user_preferences import UserPreferences
            UserPreferences.instance().set("export_last_folder", str(Path(file_path).parent))
        except Exception:
            pass

        # 3) Write the workbook (synchronous, busy cursor)
        QApplication.setOverrideCursor(QCursor(Qt.CursorShape.WaitCursor))
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            wb.remove(wb.active)
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(fill_type="solid", start_color="0078D4", end_color="0078D4")

            for sheet_name, df in sheets:
                ws = wb.create_sheet(sheet_name)
                headers = [str(c) for c in df.columns]
                ws.append(headers)
                for col_idx in range(1, len(headers) + 1):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.font = header_font
                    cell.fill = header_fill
                # Track max length per column while writing rows so we can
                # autosize without re-scanning the worksheet afterwards.
                max_lens = [len(h) for h in headers]
                for row_values in df.itertuples(index=False, name=None):
                    ws.append(list(row_values))
                    for i, v in enumerate(row_values):
                        if v is None:
                            continue
                        length = len(str(v))
                        if length > max_lens[i]:
                            max_lens[i] = length
                ws.freeze_panes = "A2"
                if headers:
                    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
                # Autosize columns: clamp to a reasonable [8, 60] character window
                for i, mlen in enumerate(max_lens):
                    width = min(max(mlen + 2, 8), 60)
                    ws.column_dimensions[get_column_letter(i + 1)].width = width

            wb.save(file_path)
        except Exception as e:
            QApplication.restoreOverrideCursor()
            logger.error(f"Global Excel export failed: {e}")
            DialogHelper.error(tr("export_all_failed"), tr("export_all_title"), dialog_parent, details=str(e))
            return
        finally:
            QApplication.restoreOverrideCursor()

        # 4) Status feedback
        msg = tr("export_all_success", sheet_count=len(sheets), path=file_path)
        try:
            w = dialog_parent.window()
            while w is not None and not hasattr(w, 'status_bar'):
                w = w.parent()
            if w is not None and hasattr(w, 'status_bar'):
                w.status_bar.set_message(msg)
        except Exception:
            pass

        if truncated_sheets:
            details = "\n".join(
                f"- {name}: {count:,} rows truncated to {EXCEL_MAX_ROWS:,}"
                for name, count in truncated_sheets
            )
            logger.warning(f"Global export — sheets truncated:\n{details}")
            DialogHelper.info(
                tr("export_all_truncated", count=len(truncated_sheets)) + "\n\n" + details,
                tr("export_all_title"), dialog_parent,
            )

    def _export_connection(self, db_conn: DatabaseConnection):
        """Export a single connection to JSON file."""
        try:
            # Ask user for file location
            default_filename = f"{db_conn.name.replace(' ', '_')}_connection.json"
            filepath, _ = QFileDialog.getSaveFileName(
                self,
                "Export Connection",
                default_filename,
                "JSON Files (*.json);;All Files (*)"
            )

            if not filepath:
                return  # User cancelled

            # Export connection
            export_data = export_connections_to_json(
                connection_ids=[db_conn.id],
                include_credentials=False  # Security: don't export passwords
            )

            # Save to file
            save_export_to_file(export_data, filepath)

            # Show success message
            summary = get_export_summary(export_data)
            DialogHelper.info(
                f"Export successful!\n\n{summary}\n\nFile: {filepath}",
                parent=self
            )

        except Exception as e:
            logger.error(f"Export failed: {e}")
            DialogHelper.error(f"Export failed: {str(e)}", parent=self)

    def _export_all_connections(self):
        """Export all connections to JSON file."""
        try:
            # Ask user for file location
            filepath, _ = QFileDialog.getSaveFileName(
                self,
                "Export All Connections",
                "connections_export.json",
                "JSON Files (*.json);;All Files (*)"
            )

            if not filepath:
                return  # User cancelled

            # Export all connections
            export_data = export_connections_to_json(
                connection_ids=None,  # All connections
                include_credentials=False  # Security: don't export passwords
            )

            # Save to file
            save_export_to_file(export_data, filepath)

            # Show success message
            summary = get_export_summary(export_data)
            DialogHelper.info(
                f"Export successful!\n\n{summary}\n\nFile: {filepath}",
                parent=self
            )

        except Exception as e:
            logger.error(f"Export failed: {e}")
            DialogHelper.error(f"Export failed: {str(e)}", parent=self)

    def _import_connections(self):
        """Import connections from JSON file."""
        try:
            # Ask user for file location
            filepath, _ = QFileDialog.getOpenFileName(
                self,
                "Import Connections",
                "",
                "JSON Files (*.json);;All Files (*)"
            )

            if not filepath:
                return  # User cancelled

            # Load import data
            import_data = load_import_from_file(filepath)

            # Check export type - must be connections or workspace (we extract connections)
            export_type = import_data.get("export_type", "")
            if export_type not in ["connections", "workspace"]:
                DialogHelper.error(
                    tr("db_export_format_unsupported", format=export_type),
                    parent=self
                )
                return

            # Import connections
            results = import_connections_from_json(import_data)

            # Refresh schema tree to show new connections
            self._refresh_schema()

            # Show results
            created = results.get("created", [])
            existing = results.get("existing", [])
            errors = results.get("errors", [])

            summary_lines = [tr("db_import_done")]
            if created:
                summary_lines.append(f"\n{tr('db_import_created', count=len(created))}")
                for name in created[:5]:
                    summary_lines.append(f"  - {name}")
                if len(created) > 5:
                    summary_lines.append(f"  ... +{len(created) - 5}")

            if existing:
                summary_lines.append(f"\n{tr('db_import_existing', count=len(existing))}")
                for name in existing[:5]:
                    summary_lines.append(f"  - {name}")
                if len(existing) > 5:
                    summary_lines.append(f"  ... +{len(existing) - 5}")

            if errors:
                summary_lines.append(f"\n{tr('db_import_errors', count=len(errors))}")
                for err in errors[:3]:
                    summary_lines.append(f"  - {err}")
                if len(errors) > 3:
                    summary_lines.append(f"  ... +{len(errors) - 3}")

            DialogHelper.info("\n".join(summary_lines), parent=self)

        except ValueError as e:
            DialogHelper.error(tr("db_import_format_error", error=str(e)), parent=self)
        except Exception as e:
            logger.error(f"Import failed: {e}")
            DialogHelper.error(tr("db_import_failed", error=str(e)), parent=self)
